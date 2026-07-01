#!/usr/bin/env python3
"""
extract_portfolio_symbols.py - Extracts stock symbols from portfolio Excel file.

Reads from temp/stocks.xlsx, table TABLE_PORTFOLIO_SUMMARY, extracts stock symbols 
from the 2nd column, and outputs to a JSON file.

Usage:
    python3 investment_screener/backend/py_services/extract_portfolio_symbols.py
"""

import json
import pandas as pd
from pathlib import Path

# Configuration
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent.parent
EXCEL_FILE = PROJECT_ROOT / "temp" / "stocks.xlsx"
OUTPUT_FILE = PROJECT_ROOT / "temp" / "portfolio_symbols.json"
TABLE_NAME = "TABLE_PORTFOLIO_SUMMARY"

# Symbols to ignore (cash positions, currency ETFs)
IGNORE_LIST = [
    "USD_CASH",
    "PSU.U",
    "DLR.U"
]


def extract_symbols():
    """Extract stock symbols from the portfolio Excel file."""
    
    if not EXCEL_FILE.exists():
        print(f"❌ Error: Excel file not found at {EXCEL_FILE}")
        return None
    
    print(f"📖 Reading {EXCEL_FILE}...")
    
    try:
        from openpyxl import load_workbook
        import warnings
        warnings.filterwarnings('ignore')
        
        wb = load_workbook(EXCEL_FILE, data_only=True)
        
        # Find the sheet containing TABLE_PORTFOLIO_SUMMARY
        target_sheet = None
        table_ref = None
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for table_name in ws.tables:
                if table_name == TABLE_NAME:
                    target_sheet = ws
                    table_ref = ws.tables[table_name].ref
                    print(f"📊 Found table '{TABLE_NAME}' in sheet '{sheet_name}' at {table_ref}")
                    break
            if target_sheet:
                break
        
        if not target_sheet:
            print(f"❌ Error: Table '{TABLE_NAME}' not found in workbook")
            return None
        
        # Parse the table range and extract data
        # table_ref is like "K4:AA38"
        from openpyxl.utils import range_boundaries
        min_col, min_row, max_col, max_row = range_boundaries(table_ref)
        
        # Read all rows from the table (skip header row)
        symbols = []
        for row in range(min_row + 1, max_row + 1):  # +1 to skip header
            # 2nd column of the table (B column relative to table start)
            symbol_cell = target_sheet.cell(row=row, column=min_col + 1)  # +1 for 2nd column
            if symbol_cell.value:
                sym = str(symbol_cell.value).strip()
                if sym and sym not in IGNORE_LIST and sym.upper() not in IGNORE_LIST:
                    # Skip if it looks like a header
                    if sym.upper() not in ["SYMBOL", "TICKER", "STOCK"]:
                        symbols.append(sym)
        
        # Remove duplicates while preserving order
        seen = set()
        unique_symbols = []
        for sym in symbols:
            if sym not in seen:
                seen.add(sym)
                unique_symbols.append(sym)
        
        print(f"✅ Extracted {len(unique_symbols)} unique symbols")
        print(f"❌ Ignored: {IGNORE_LIST}")
        
        return unique_symbols
        
    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")
        import traceback
        traceback.print_exc()
        return None


def main():
    symbols = extract_symbols()
    
    if symbols is None:
        return
    
    # Output to JSON
    output_data = {
        "source": str(EXCEL_FILE),
        "table": TABLE_NAME,
        "ignored": IGNORE_LIST,
        "count": len(symbols),
        "symbols": symbols
    }
    
    with open(OUTPUT_FILE, "w") as f:
        json.dump(output_data, f, indent=2)
    
    print(f"\n📁 Output saved to: {OUTPUT_FILE}")
    print(f"\n📋 Symbols extracted:")
    for i, sym in enumerate(symbols, 1):
        print(f"   {i:2d}. {sym}")


if __name__ == "__main__":
    main()
